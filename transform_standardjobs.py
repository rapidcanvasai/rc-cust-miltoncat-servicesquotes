"""
Transform 'Standard Job Pricing.xlsx' into standardJobs.json
for the Milton CAT Service Quotes frontend.

Column mapping (from Excel header row):
  A (0): Standard Job ID
  B (1): Make
  C (2): Model
  D (3): From Serial  -> serial prefix = first 3 chars
  E (4): To Serial
  F (5): Job Code
  G (6): Job Code Desc
  H (7): Comp Code
  I (8): Comp Desc
  J (9): Qty Code
  K (10): Qty Code Desc
  L (11): Modifier Code
  M (12): Modifier Code Desc
  N (13): Parts Price
  O (14): Labor Price
  P (15): Misc Price
  Q (16): Shop Field
  V (21): Estimated Duration

Filters:
  - Discards rows where parts, labor, and misc are all zero
    (per Tim Dailey: never flat-rate priced, unreliable)

Output key format: model|serialPrefix|jobCode|compCode
"""

import json
import openpyxl
from collections import defaultdict

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_PATH = "usecases/servicesquotes/Standard Job Pricing.xlsx"
OUTPUT_PATH = "usecases/servicesquotes/frontend/src/data/standardJobs.json"

# Column indices (0-based)
COL_STJ_ID = 0
COL_MAKE = 1
COL_MODEL = 2
COL_FROM_SERIAL = 3
COL_TO_SERIAL = 4
COL_JOB_CODE = 5
COL_JOB_DESC = 6
COL_COMP_CODE = 7
COL_COMP_DESC = 8
COL_MODIFIER_CODE = 11
COL_MODIFIER_DESC = 12
COL_PARTS = 13
COL_LABOR = 14
COL_MISC = 15
COL_SHOP_FIELD = 16
COL_DURATION = 21


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def classify_location(shop_field: str) -> str:
    """S = Shop, F = Field based on the Shop/Field column value."""
    return "S" if shop_field.lower().startswith("s") else "F"


def main():
    print(f"Reading {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Collectors
    jobs: dict[str, list[dict]] = defaultdict(list)
    model_set: dict[str, str] = {}
    job_code_set: dict[str, str] = {}
    comp_code_set: dict[str, str] = {}
    serial_prefix_set: set[str] = set()

    total = 0
    skipped_zero = 0
    skipped_missing = 0
    included = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= 1:
            continue  # skip title row + header row
        total += 1

        vals = list(row)

        # Extract fields
        stj_id = safe_str(vals[COL_STJ_ID])
        make = safe_str(vals[COL_MAKE])
        model = safe_str(vals[COL_MODEL])
        from_serial = safe_str(vals[COL_FROM_SERIAL])
        job_code = safe_str(vals[COL_JOB_CODE])
        job_desc = safe_str(vals[COL_JOB_DESC])
        comp_code = safe_str(vals[COL_COMP_CODE])
        comp_desc = safe_str(vals[COL_COMP_DESC])
        modifier_desc = safe_str(vals[COL_MODIFIER_DESC])
        shop_field = safe_str(vals[COL_SHOP_FIELD])

        parts = safe_float(vals[COL_PARTS])
        labor = safe_float(vals[COL_LABOR])
        misc = safe_float(vals[COL_MISC])
        duration = safe_float(vals[COL_DURATION])

        # Must have model, job code, and comp code
        if not model or not job_code or not comp_code:
            skipped_missing += 1
            continue

        # Filter: discard zero-value records (parts=0, labor=0, misc=0)
        if parts == 0 and labor == 0 and misc == 0:
            skipped_zero += 1
            continue

        # Serial prefix: first 3 characters of From Serial
        serial_prefix = from_serial[:3] if len(from_serial) >= 3 else ""

        # Build key: Model|SerialPrefix|JobCode|CompCode
        key = f"{model}|{serial_prefix}|{job_code}|{comp_code}"

        # Build entry
        entry: dict = {
            "s": stj_id,
            "p": round(parts, 2),
            "l": round(labor, 2),
            "m": round(misc, 2),
            "d": round(duration, 1),
            "sf": classify_location(shop_field),
        }
        if serial_prefix:
            entry["sp"] = serial_prefix
        if modifier_desc:
            entry["md"] = modifier_desc

        jobs[key].append(entry)

        # Collect unique values
        if model not in model_set:
            model_set[model] = make or "CAT"
        if job_code not in job_code_set:
            job_code_set[job_code] = job_desc
        if comp_code not in comp_code_set:
            comp_code_set[comp_code] = comp_desc
        if serial_prefix:
            serial_prefix_set.add(serial_prefix)

        included += 1

    wb.close()

    # Build output structure
    output = {
        "hasSerialPrefix": True,
        "models": sorted(
            [{"id": mid, "make": make} for mid, make in model_set.items()],
            key=lambda x: x["id"],
        ),
        "serialPrefixes": sorted(
            [{"prefix": p} for p in serial_prefix_set],
            key=lambda x: x["prefix"],
        ),
        "jobCodes": sorted(
            [{"code": c, "desc": d} for c, d in job_code_set.items()],
            key=lambda x: x["code"],
        ),
        "compCodes": sorted(
            [{"code": c, "desc": d} for c, d in comp_code_set.items()],
            key=lambda x: x["code"],
        ),
        "jobs": dict(sorted(jobs.items())),
    }

    # Write JSON
    print(f"\nWriting {OUTPUT_PATH} ...")
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(",", ":"))

    # Stats
    total_entries = sum(len(v) for v in jobs.values())
    print(f"\n{'='*50}")
    print(f"TRANSFORMATION COMPLETE")
    print(f"{'='*50}")
    print(f"Total rows read:        {total:>10,}")
    print(f"Skipped (missing data): {skipped_missing:>10,}")
    print(f"Skipped (all zeros):    {skipped_zero:>10,}")
    print(f"Included:               {included:>10,}")
    print(f"{'='*50}")
    print(f"Unique models:          {len(model_set):>10,}")
    print(f"Unique serial prefixes: {len(serial_prefix_set):>10,}")
    print(f"Unique job codes:       {len(job_code_set):>10,}")
    print(f"Unique component codes: {len(comp_code_set):>10,}")
    print(f"Unique combinations:    {len(jobs):>10,}")
    print(f"Total entries:          {total_entries:>10,}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
