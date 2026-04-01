"""
Transform additional parts Excel files and merge into partsData.json.

Reads parts transaction data from data/additional_parts_data/*.xlsx,
maps each transaction to a combo key via workOrders.json segment lookup,
and merges with the existing partsData.json.

Join logic:
  Parts Excel Project ID "CSR0652648_01"
    → strip "CSR" → segment_id "0652648_01"
    → lookup in workOrders.json → combo_key "G125|NGG|540|1350"
    → add part entry to partsData[combo_key]
"""

import json
import os
import glob
import openpyxl
from collections import defaultdict

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
ADDITIONAL_PARTS_DIR = "data/additional_parts_data"
WORK_ORDERS_PATH = "frontend/src/data/workOrders.json"
PARTS_DATA_PATH = "frontend/src/data/partsData.json"
OUTPUT_PATH = "frontend/src/data/partsData.json"

# Column indices (0-based, row 0 = header)
COL_PROJECT_ID = 1
COL_ITEM_NUMBER = 3
COL_QUANTITY = 7
COL_COST_PRICE = 8
COL_SALES_PRICE = 10
COL_LINE_CATEGORY_GROUP = 12


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def build_segment_lookup(wo_path: str) -> dict[str, str]:
    """Build segment_id -> combo_key mapping from workOrders.json."""
    with open(wo_path, "r", encoding="utf-8") as f:
        wo_data = json.load(f)

    lookup = {}
    for combo_key, entries in wo_data["jobs"].items():
        for entry in entries:
            lookup[entry["s"]] = combo_key

    print(f"Built segment lookup: {len(lookup):,} segment IDs")
    return lookup


def load_existing_parts(path: str) -> dict[str, list[dict]]:
    """Load existing partsData.json."""
    if not os.path.exists(path):
        print(f"No existing partsData.json found, starting fresh")
        return {}

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    total_entries = sum(len(v) for v in data.values())
    print(f"Loaded existing partsData: {len(data):,} combo keys, {total_entries:,} entries")
    return data


def process_excel_file(
    filepath: str,
    segment_lookup: dict[str, str],
) -> tuple[dict[str, list[dict]], int, int, int]:
    """Process one parts Excel file. Returns (parts_by_key, total, matched, unmatched)."""
    print(f"\nReading {os.path.basename(filepath)} ...")
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    new_parts: dict[str, list[dict]] = defaultdict(list)
    total = 0
    matched = 0
    unmatched = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        total += 1

        vals = list(row)

        project_id = safe_str(vals[COL_PROJECT_ID])
        if not project_id:
            unmatched += 1
            continue

        segment_id = project_id.replace("CSR", "", 1)
        combo_key = segment_lookup.get(segment_id)
        if combo_key is None:
            unmatched += 1
            continue

        part_number = safe_str(vals[COL_ITEM_NUMBER])
        if not part_number:
            unmatched += 1
            continue

        qty = safe_float(vals[COL_QUANTITY])
        qty = int(qty) if qty == int(qty) else qty

        entry = {
            "pn": part_number,
            "qty": qty,
            "cost": round(safe_float(vals[COL_COST_PRICE]), 2),
            "price": round(safe_float(vals[COL_SALES_PRICE]), 2),
            "cat": safe_str(vals[COL_LINE_CATEGORY_GROUP]),
        }

        new_parts[combo_key].append(entry)
        matched += 1

    wb.close()

    print(f"  Total rows: {total:,}")
    print(f"  Matched: {matched:,}")
    print(f"  Unmatched: {unmatched:,}")

    return dict(new_parts), total, matched, unmatched


def main():
    print("=" * 50)
    print("Parts Data Transformation")
    print("=" * 50)

    segment_lookup = build_segment_lookup(WORK_ORDERS_PATH)
    existing_parts = load_existing_parts(PARTS_DATA_PATH)
    existing_count = sum(len(v) for v in existing_parts.values())

    excel_files = sorted(glob.glob(os.path.join(ADDITIONAL_PARTS_DIR, "*.xlsx")))
    if not excel_files:
        print(f"No .xlsx files found in {ADDITIONAL_PARTS_DIR}")
        return

    print(f"\nFound {len(excel_files)} Excel file(s)")

    all_new_parts: dict[str, list[dict]] = defaultdict(list)
    grand_total = 0
    grand_matched = 0
    grand_unmatched = 0

    for filepath in excel_files:
        new_parts, total, matched, unmatched = process_excel_file(filepath, segment_lookup)
        grand_total += total
        grand_matched += matched
        grand_unmatched += unmatched
        for key, entries in new_parts.items():
            all_new_parts[key].extend(entries)

    # Merge
    print(f"\nMerging ...")
    merged = dict(existing_parts)
    new_keys = 0
    for combo_key, entries in all_new_parts.items():
        if combo_key in merged:
            merged[combo_key].extend(entries)
        else:
            merged[combo_key] = entries
            new_keys += 1

    sorted_merged = dict(sorted(merged.items()))

    print(f"Writing {OUTPUT_PATH} ...")
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(sorted_merged, f, ensure_ascii=False, separators=(",", ":"))

    final_count = sum(len(v) for v in sorted_merged.values())
    print(f"\n{'=' * 50}")
    print(f"COMPLETE")
    print(f"{'=' * 50}")
    print(f"Excel rows read:        {grand_total:>10,}")
    print(f"Matched to combo key:   {grand_matched:>10,}")
    print(f"Unmatched:              {grand_unmatched:>10,}")
    print(f"{'=' * 50}")
    print(f"Existing combo keys:    {len(existing_parts):>10,}")
    print(f"Existing entries:       {existing_count:>10,}")
    print(f"Final combo keys:       {len(sorted_merged):>10,}")
    print(f"Final entries:          {final_count:>10,}")
    print(f"New combo keys:         {new_keys:>10,}")
    print(f"New entries added:      {final_count - existing_count:>10,}")
    print(f"{'=' * 50}")


if __name__ == "__main__":
    main()
