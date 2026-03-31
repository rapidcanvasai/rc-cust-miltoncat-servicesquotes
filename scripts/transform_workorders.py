"""
Transform 'Service Call Job and Component Code.xlsx' into workOrders.json
for the Milton CAT Service Quotes frontend.

Filters:
  - Only 'Closed' segments
  - Must have both Job Code AND Component Code
  - Discards cancelled and incomplete rows (per Tim Dailey's instructions)

Output mirrors standardJobs.json structure (4-part key with serial prefix):
{
  "hasSerialPrefix": true,
  "models": [{ "id": "325", "make": "CAT" }],
  "serialPrefixes": [{ "prefix": "GWR" }],
  "jobCodes": [{ "code": "010", "desc": "REMOVE AND INSTALL" }],
  "compCodes": [{ "code": "1000", "desc": "ENGINE" }],
  "jobs": {
    "325|GWR|010|1000": [
      { "s": "0975465_05", "p": 369.57, "l": 657.79, "m": 156.8, "d": 6, "sf": "S", "sp": "GWR" }
    ]
  }
}
"""

import json
import openpyxl
from collections import defaultdict

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_PATH = "usecases/servicesquotes/Service Call Job and Component Code.xlsx"
OUTPUT_PATH = "usecases/servicesquotes/frontend/src/data/workOrders.json"

# Cost centers that map to Shop (S) vs Field (F)
SHOP_COST_CENTERS = {
    "SHG", "SHE", "SHM", "SHR", "SHC", "SHH",  # Shop variants
    "PSC", "PSG", "PSR",                          # Parts/Service centers
    "UND", "PVM", "WLD", "IND", "MPC",            # Internal shop operations
}
# Everything else defaults to Field (F)

# Column indices (0-based, from row 2 = header row)
COL_DIVISION = 0
COL_STORE = 1
COL_COST_CENTER = 2
COL_SERVICE_CALL = 3
COL_SEGMENT_ID = 4
COL_MODEL = 5
COL_SERIAL = 6
COL_STJ = 7
COL_SEQ = 8
COL_JOB_CODE = 9
COL_JOB_DESC = 10
COL_COMPONENT = 11
COL_COMP_DESC = 12
COL_MODIFIER_CODE = 13
COL_MODIFIER_DESC = 14
COL_QTY_CODE = 15
COL_QTY_DESC = 16
COL_PRICING_QTY = 17
COL_STATUS = 18
COL_CREATED = 19
COL_FIRST_INVOICE = 20
COL_LAST_INVOICE = 21
COL_EMPLOYEE = 22
COL_EMPLOYEE_NUM = 23
COL_SEGMENT_AMT = 24
COL_INVOICED_AMT = 25
COL_INVOICED_HOURS = 26
COL_PART_ACTUAL = 27
COL_PART_FLAT = 28
COL_LABOR_ACTUAL = 29
COL_LABOR_FLAT = 30
COL_MISC_ACTUAL = 31
COL_MISC_FLAT = 32


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def classify_location(cost_center: str) -> str:
    """S = Shop, F = Field based on cost center code."""
    return "S" if cost_center.upper() in SHOP_COST_CENTERS else "F"


def main():
    print(f"Reading {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Collectors
    jobs: dict[str, list[dict]] = defaultdict(list)
    model_set: dict[str, str] = {}       # id -> make (placeholder)
    job_code_set: dict[str, str] = {}    # code -> desc
    comp_code_set: dict[str, str] = {}   # code -> desc
    serial_prefix_set: set[str] = set()

    total = 0
    skipped_status = 0
    skipped_missing = 0
    included = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= 1:
            continue  # skip title row + header row
        total += 1

        vals = list(row)

        # Filter: only Closed
        status = safe_str(vals[COL_STATUS])
        if status != "Closed":
            skipped_status += 1
            continue

        # Filter: must have both job code and component code
        job_code = safe_str(vals[COL_JOB_CODE])
        component = safe_str(vals[COL_COMPONENT])
        if not job_code or not component:
            skipped_missing += 1
            continue

        # Extract fields
        model = safe_str(vals[COL_MODEL])
        serial_raw = safe_str(vals[COL_SERIAL])
        serial_prefix = serial_raw[:3] if len(serial_raw) >= 3 else ""
        segment_id = safe_str(vals[COL_SEGMENT_ID])
        job_desc = safe_str(vals[COL_JOB_DESC])
        comp_desc = safe_str(vals[COL_COMP_DESC])
        cost_center = safe_str(vals[COL_COST_CENTER])
        modifier_desc = safe_str(vals[COL_MODIFIER_DESC])

        part_actual = safe_float(vals[COL_PART_ACTUAL])
        part_flat = safe_float(vals[COL_PART_FLAT])
        labor_actual = safe_float(vals[COL_LABOR_ACTUAL])
        labor_flat = safe_float(vals[COL_LABOR_FLAT])
        misc_actual = safe_float(vals[COL_MISC_ACTUAL])
        misc_flat = safe_float(vals[COL_MISC_FLAT])
        hours = safe_float(vals[COL_INVOICED_HOURS])

        # Use flat rate if available, otherwise actual
        parts = part_flat if part_flat > 0 else part_actual
        labor = labor_flat if labor_flat > 0 else labor_actual
        misc = misc_flat if misc_flat > 0 else misc_actual

        # Skip zero-value rows (no real work done)
        if parts == 0 and labor == 0 and misc == 0 and hours == 0:
            skipped_missing += 1
            continue

        # Build key: Model|SerialPrefix|JobCode|Component
        key = f"{model}|{serial_prefix}|{job_code}|{component}"

        # Build entry (mirrors STJ structure)
        # Strip CSR prefix from segment ID to save space (18MB -> smaller)
        short_id = segment_id.replace("CSR", "", 1)

        # Extract date for recency weighting (prefer last invoice, then first invoice, then created)
        raw_date = vals[COL_LAST_INVOICE] or vals[COL_FIRST_INVOICE] or vals[COL_CREATED]
        date_str = ""
        if raw_date:
            from datetime import datetime
            if isinstance(raw_date, datetime):
                date_str = raw_date.strftime("%Y-%m")
            else:
                try:
                    parsed_date = datetime.strptime(str(raw_date).strip()[:10], "%Y-%m-%d")
                    date_str = parsed_date.strftime("%Y-%m")
                except (ValueError, TypeError):
                    pass

        entry: dict = {
            "s": short_id,
            "p": round(parts, 2),
            "l": round(labor, 2),
            "m": round(misc, 2),
            "d": round(hours, 1),
            "sf": classify_location(cost_center),
        }
        if date_str:
            entry["dt"] = date_str
        if serial_prefix:
            entry["sp"] = serial_prefix
        if modifier_desc:
            entry["md"] = modifier_desc

        jobs[key].append(entry)

        # Collect unique models, job codes, component codes
        if model not in model_set:
            model_set[model] = "CAT"
        if job_code not in job_code_set:
            job_code_set[job_code] = job_desc
        if component not in comp_code_set:
            comp_code_set[component] = comp_desc
        if serial_prefix:
            serial_prefix_set.add(serial_prefix)

        included += 1

    wb.close()

    # Build output structure
    output = {
        "hasSerialPrefix": True,
        "models": sorted(
            [{"id": mid, "make": make} for mid, make in model_set.items()],
            key=lambda x: x["id"],
        ),
        "serialPrefixes": sorted(
            [{"prefix": p} for p in serial_prefix_set],
            key=lambda x: x["prefix"],
        ),
        "jobCodes": sorted(
            [{"code": c, "desc": d} for c, d in job_code_set.items()],
            key=lambda x: x["code"],
        ),
        "compCodes": sorted(
            [{"code": c, "desc": d} for c, d in comp_code_set.items()],
            key=lambda x: x["code"],
        ),
        "jobs": dict(sorted(jobs.items())),
    }

    # Write JSON
    print(f"\nWriting {OUTPUT_PATH} ...")
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(",", ":"))

    # Stats
    total_entries = sum(len(v) for v in jobs.values())
    print(f"\n{'='*50}")
    print(f"TRANSFORMATION COMPLETE")
    print(f"{'='*50}")
    print(f"Total rows read:        {total:>10,}")
    print(f"Skipped (not Closed):   {skipped_status:>10,}")
    print(f"Skipped (missing data): {skipped_missing:>10,}")
    print(f"Included:               {included:>10,}")
    print(f"{'='*50}")
    print(f"Unique models:          {len(model_set):>10,}")
    print(f"Unique serial prefixes: {len(serial_prefix_set):>10,}")
    print(f"Unique job codes:       {len(job_code_set):>10,}")
    print(f"Unique component codes: {len(comp_code_set):>10,}")
    print(f"Unique combinations:    {len(jobs):>10,}")
    print(f"Total entries:          {total_entries:>10,}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
